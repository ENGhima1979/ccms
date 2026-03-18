"""
مساعدات النظام — تقارير PDF، Excel، QR، منطق الصلاحيات
"""
import datetime, io, json, os, uuid
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from flask import session
from models import get_db
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# -- Auth helpers --------------------------------------
def get_user_project_ids():
    if session.get('role') in ('super_admin','admin') or session.get('all_projects'):
        return None
    uid = session.get('user_id')
    if not uid: return []
    conn = get_db()
    rows = conn.execute("SELECT project_id FROM user_projects WHERE user_id=?", (uid,)).fetchall()
    conn.close()
    return [r['project_id'] for r in rows]

def get_visible_projects(conn=None):
    close = conn is None
    if close: conn = get_db()
    pid_list = get_user_project_ids()
    if pid_list is None:
        rows = conn.execute("SELECT * FROM projects WHERE is_active=1 ORDER BY name").fetchall()
    elif len(pid_list) == 0:
        rows = []
    else:
        ph = ','.join('?'*len(pid_list))
        rows = conn.execute(f"SELECT * FROM projects WHERE id IN ({ph}) AND is_active=1 ORDER BY name", pid_list).fetchall()
    if close: conn.close()
    return rows

def apply_project_filter(sql, params, pid_list, alias='c'):
    if pid_list is None: return sql, params
    if len(pid_list) == 0: return sql + " AND 1=0", params
    ph = ','.join('?'*len(pid_list))
    return sql + f" AND {alias}.project_id IN ({ph})", params + pid_list

def can_delete():
    return session.get('role') in ('super_admin', 'admin')

def can_manage_users():
    return session.get('role') in ('super_admin', 'admin')

def can_manage_projects():
    return session.get('role') in ('super_admin', 'admin', 'manager')

def get_unread_count():
    uid = session.get('user_id')
    if not uid: return 0
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) as c FROM notifications WHERE user_id=? AND is_read=0", (uid,)).fetchone()['c']
    conn.close()
    return n

def get_pending_workflow_count():
    uid = session.get('user_id')
    if not uid: return 0
    conn = get_db()
    n = conn.execute("""SELECT COUNT(*) as c FROM workflow_steps ws
                        JOIN correspondence c ON ws.correspondence_id=c.id
                        WHERE ws.assigned_to=? AND ws.status='pending'
                        AND c.is_deleted=0""", (uid,)).fetchone()['c']
    conn.close()
    return n

def create_notification(user_id, type_, title, body=None, link=None, conn=None):
    close = conn is None
    if close: conn = get_db()
    conn.execute("""INSERT INTO notifications (id,user_id,type,title,body,link,created_at)
                    VALUES (?,?,?,?,?,?,?)""",
                 (str(uuid.uuid4()), user_id, type_, title, body, link,
                  datetime.datetime.now().isoformat(timespec='seconds')))
    if close: conn.commit(); conn.close()

# -- PDF Generation ------------------------------------
def generate_letter_pdf(corr, company, attachments=None, include_letterhead=True):
    """Generate professional Arabic PDF with company logo"""
    import os, sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from arabic_utils import arabic_text as ar
    buffer = io.BytesIO()

    BASE = os.path.dirname(os.path.abspath(__file__))

    font_reg  = os.path.join(BASE, 'fonts', 'DejaVuSans.ttf')
    font_bold = os.path.join(BASE, 'fonts', 'DejaVuSans-Bold.ttf')
    try:
        if os.path.exists(font_reg):
            pdfmetrics.registerFont(TTFont('ArabicFont', font_reg))
        if os.path.exists(font_bold):
            pdfmetrics.registerFont(TTFont('ArabicFontBold', font_bold))
        af, afb = 'ArabicFont', 'ArabicFontBold'
    except:
        af, afb = 'Helvetica', 'Helvetica-Bold'

    primary   = colors.HexColor(company.get('primary_color','#00b4d8'))
    light_bg  = colors.HexColor('#f0f8ff')
    dark_text = colors.HexColor('#0d1b2a')

    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2.5*cm, bottomMargin=2*cm)

    def S(name, font=None, size=11, align=TA_RIGHT, color=None, bold=False, leading=22, space=4):
        return ParagraphStyle(name,
            fontName=font or (afb if bold else af),
            fontSize=size, alignment=align,
            textColor=color or dark_text,
            leading=leading, spaceAfter=space, wordWrap='RTL')

    s_body   = S('body', size=12, leading=24)
    s_small  = S('small', size=10, color=colors.HexColor('#555555'), leading=16)
    s_h1     = S('h1', size=16, bold=True, color=primary, align=TA_CENTER, leading=24)
    s_ref    = S('ref', size=10, color=colors.grey, leading=16)
    s_subj   = S('subj', size=13, bold=True, color=colors.HexColor('#023e8a'), leading=22)
    s_sig    = S('sig', size=11, bold=True, leading=20)

    story = []

    co_name   = company.get('name','')
    co_phone  = company.get('phone','')
    co_email  = company.get('email','')
    co_addr   = company.get('address','')
    co_cr     = company.get('cr_number','')
    logo_path = company.get('logo_path','')

    # --- HEADER with LOGO ------------------------------------------
    # بناء رأس الصفحة: شعار + اسم الشركة (يسار) | بيانات الاتصال (يمين)

    # الشعار: نحدد أبعاده أولاً قبل إدراجه في الجدول
    logo_img = None
    if logo_path:
        possible_paths = [
            os.path.join(BASE, 'instance', 'uploads', logo_path),
            os.path.join(BASE, 'static', logo_path),
            os.path.join(BASE, logo_path),
        ]
        for lp in possible_paths:
            if os.path.exists(lp):
                try:
                    from reportlab.platypus import Image as RLImage
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(lp)
                    orig_w, orig_h = pil_img.size
                    logo_w = 2.5 * cm
                    logo_h = logo_w * (orig_h / orig_w)
                    if logo_h > 2.2 * cm:
                        logo_h = 2.2 * cm
                        logo_w = logo_h * (orig_w / orig_h)
                    logo_img = RLImage(lp, width=logo_w, height=logo_h)
                    logo_img.hAlign = 'RIGHT'
                except Exception:
                    logo_img = None
                break

    # اسم الشركة
    co_name_para = Paragraph(
        f"<b>{ar(co_name)}</b>",
        ParagraphStyle('hl', fontName=afb, fontSize=16, alignment=TA_RIGHT,
                       textColor=primary, leading=22, spaceAfter=0))

    # بيانات الاتصال
    hdr_right_lines = []
    if co_phone: hdr_right_lines.append(f"{ar('هاتف')}: {co_phone}")
    if co_email: hdr_right_lines.append(f"{ar('بريد')}: {co_email}")
    if co_addr:  hdr_right_lines.append(ar(co_addr))
    if co_cr:    hdr_right_lines.append(f"{ar('س.ت')}: {co_cr}")
    hdr_right = Paragraph(
        "<br/>".join(hdr_right_lines) if hdr_right_lines else "",
        ParagraphStyle('hr', fontName=af, fontSize=9, alignment=TA_RIGHT,
                       textColor=colors.grey, leading=15))

    # بناء خلية الشعار + الاسم بدون KeepTogether
    if logo_img:
        # جدول داخلي: شعار فوق الاسم
        inner = Table(
            [[logo_img], [co_name_para]],
            colWidths=[9*cm],
            rowHeights=None
        )
        inner.setStyle(TableStyle([
            ('ALIGN',  (0,0),(-1,-1),'RIGHT'),
            ('VALIGN', (0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',    (0,0),(-1,-1), 2),
            ('BOTTOMPADDING', (0,0),(-1,-1), 2),
        ]))
        logo_cell = inner
    else:
        logo_cell = co_name_para

    hdr = Table([[logo_cell, hdr_right]], colWidths=[9*cm, 8.4*cm])
    hdr.setStyle(TableStyle([
        ('VALIGN',        (0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',         (0,0),(-1,-1),'RIGHT'),
        ('LINEBELOW',     (0,0),(-1,-1), 2.5, primary),
        ('BOTTOMPADDING', (0,0),(-1,-1), 10),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 0.4*cm))

    # --- REFERENCE BAR ----------------------------------
    priority_map = {'urgent':'⚡ عاجل','high':'▲ عالية','normal':'● عادية','low':'▽ منخفضة'}
    type_map     = {'out':'صادر','in':'وارد','internal':'داخلي'}

    ref_data = [[
        Paragraph(f"<b>{ar('رقم المرجع')}</b><br/>{corr.get('ref_num','')}", s_ref),
        Paragraph(f"<b>{ar('التاريخ')}</b><br/>{corr.get('date','')}", s_ref),
        Paragraph(f"<b>{ar('النوع')}</b><br/>{ar(type_map.get(corr.get('type','out'),'صادر'))}", s_ref),
        Paragraph(f"<b>{ar('الأولوية')}</b><br/>{ar(priority_map.get(corr.get('priority','normal'),'عادية'))}", s_ref),
    ]]
    if corr.get('proj_name'):
        ref_data[0].append(Paragraph(f"<b>{ar('المشروع')}</b><br/>{ar(corr['proj_name'])}", s_ref))
        ref_t = Table(ref_data, colWidths=[4.3*cm, 3.2*cm, 2.8*cm, 2.8*cm, 4.3*cm])
    else:
        ref_t = Table(ref_data, colWidths=[4.5*cm, 3.5*cm, 3.5*cm, 5.3*cm])

    ref_t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), light_bg),
        ('BOX',(0,0),(-1,-1), 0.5, colors.HexColor('#90c7e8')),
        ('GRID',(0,0),(-1,-1), 0.3, colors.HexColor('#c8e6f0')),
        ('PADDING',(0,0),(-1,-1), 8),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(ref_t)
    story.append(Spacer(1, 0.5*cm))

    # --- RECIPIENT --------------------------------------
    story.append(Paragraph(f"{ar('السادة')} / <b>{ar(corr.get('party',''))}</b>", s_body))
    story.append(Paragraph(ar('المحترمين'), s_body))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(ar('السلام عليكم ورحمة الله وبركاته،،،'), s_body))
    story.append(Spacer(1, 0.3*cm))

    # --- SUBJECT ----------------------------------------
    story.append(HRFlowable(width="100%", thickness=0.5, color=primary))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"{ar('الموضوع')}: <b>{ar(corr.get('subject',''))}</b>", s_subj))
    story.append(Spacer(1, 0.2*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=primary))
    story.append(Spacer(1, 0.5*cm))

    # --- BODY --------------------------------------------
    body_text = corr.get('body','') or ''
    for para in body_text.split('\n'):
        if para.strip():
            story.append(Paragraph(ar(para.strip()), s_body))
            story.append(Spacer(1, 0.15*cm))

    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph(ar('وتفضلوا بقبول فائق الاحترام والتقدير،،،'), s_body))
    story.append(Spacer(1, 1.2*cm))

    # --- SIGNATURE --------------------------------------
    sig_lines = [ar(co_name)]
    if corr.get('sender_name'):
        sig_lines.append(ar(corr['sender_name']))
        if corr.get('sender_title'):
            sig_lines.append(ar(corr['sender_title']))
    
    sig_data = [[Paragraph("<br/>".join(sig_lines), s_sig)]]
    sig_t = Table(sig_data, colWidths=[8*cm], hAlign='RIGHT')
    sig_t.setStyle(TableStyle([
        ('LINEABOVE',(0,0),(-1,-1), 1.5, primary),
        ('TOPPADDING',(0,0),(-1,-1), 10),
        ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ]))
    story.append(sig_t)

    # --- ATTACHMENTS -------------------------------------
    if attachments:
        story.append(Spacer(1, 0.8*cm))
        att_rows = [[Paragraph(f"<b>المرفقات ({len(attachments)})</b>",
                    ParagraphStyle('at', fontName=afb, fontSize=11, textColor=primary, alignment=TA_RIGHT))]]
        for a in attachments:
            att_rows.append([Paragraph(f"• {a.get('filename','')}", s_small)])
        att_t = Table(att_rows, colWidths=[17.4*cm])
        att_t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#f8f9fa')),
            ('BOX',(0,0),(-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('PADDING',(0,0),(-1,-1), 8),
        ]))
        story.append(att_t)

    # --- FOOTER LINE -------------------------------------
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.3, color=colors.lightgrey))
    story.append(Paragraph(
        f"هذه المراسلة سرية وموجهة للمعنيين فقط  |  {co_name}  |  {corr.get('ref_num','')}",
        ParagraphStyle('ft', fontName=af, fontSize=8, alignment=TA_CENTER,
                       textColor=colors.grey, leading=14)))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()

def generate_excel_report(rows, columns, title='تقرير', sheet_name='تقرير'):
    """Generate professional Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    # Header row
    header_fill = PatternFill("solid", fgColor="0077b6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='90c7e8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col.get('width', 20)

    # Data rows
    alt_fill = PatternFill("solid", fgColor="E8F4FD")
    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, col in enumerate(columns, 1):
            val = row.get(col['key'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else '')
            cell.fill = fill
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border

    ws.row_dimensions[1].height = 30

    # Title row above headers
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.fill = PatternFill("solid", fgColor="023e8a")
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_qr_svg(data):
    """Generate QR code SVG as base64"""
    from qr_simple import generate_qr_b64
    return generate_qr_b64(data)
